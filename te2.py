import codecs
import csv
import os
import re
from typing import List, Any

import openpyxl
import xlrd
from datetime import date,datetime

import sys, os
if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']
from PyQt5.QtWidgets import QMainWindow, QTextEdit, QApplication
from PyQt5.QtCore import QTimer

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QFont, QPalette

import sys
#PyQt5中使用的基本控件都在PyQt5.QtWidgets模块中
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
#导入designer工具生成的login模块



cols_time = []
cols_text = []
cols_time_s = []
cols_text_s = []

lists ={'共产党'}
def read_excel(first_name):

    url = first_name
    # 打开文件
    workbook = xlrd.open_workbook(url)
    # workbook = xlrd.open_workbook(r'C:\Users\hasee\Desktop\WFTV-1(白天)2019-11-16.xls')

    # 根据sheet索引或者名称获取sheet内容
    sheet = workbook.sheet_by_index(0) # sheet索引从0开始
    # sheet = workbook.sheet_by_name('Sheet1')

    # sheet的名称，行数，列数
    # print (sheet.name,sheet.nrows,sheet.ncols)

    cols_time = sheet.col_values(3)  # 获取第列内容
    cols_text = sheet.col_values(5)  # 获取第列内容

    # 格式变化
    if cols_time[0] != '播出时间' and cols_text[0] != '名称' :
        cols_time = []
        cols_text = []
        cols_time = sheet.col_values(2)  # 获取第列内容
        cols_text = sheet.col_values(4)  # 获取第列内容

    cols_time.pop(0)  #去“时间”
    cols_text.pop(0)  #去“名称”



    for index in range(len(cols_text)):

        pd_test1 = re.findall('HD测试卡1分钟',cols_text[index])
        if pd_test1 :
            cols_text[index] = ''

        #统一广告格式
        pd1 = re.findall('广告*', cols_text[index])
        if  pd1  :
            cols_text[index] = '广告'

        pd2 = re.findall('公益*', cols_text[index])
        if  pd2  :
            cols_text[index] = '公益广告'


        #去掉周*（后期抽象一个方法）
        pd_Monday = re.search('周一', cols_text[index])
        if pd_Monday != None:
            cols_text[index] = re.sub('周一(.+|.*)','',cols_text[index])

        pd_Tuesday = re.search('周二', cols_text[index])
        if pd_Tuesday != None:
            cols_text[index] = re.sub('周二(.+|.*)','',cols_text[index])

        pd_Wednesday = re.search('周三', cols_text[index])
        if pd_Wednesday != None:
            cols_text[index] = re.sub('周三(.+|.*)','',cols_text[index])


        pd_Thursday = re.search('周四', cols_text[index])
        if pd_Thursday != None:
            cols_text[index] = re.sub('周四(.+|.*)','',cols_text[index])

        pd_Friday = re.search('周五', cols_text[index])
        if pd_Friday != None:
           cols_text[index] = re.sub('(周五)(.+|.*)','',cols_text[index])

        pd_Saturday = re.search('周六', cols_text[index])
        if pd_Saturday != None:
            cols_text[index] = re.sub('周六(.+|.*)','',cols_text[index])

        pd_Sunday = re.search('周日', cols_text[index])
        if pd_Sunday != None:
            cols_text[index] = re.sub('周日(.+|.*)','',cols_text[index])

        #去掉片头片尾
        pd_titles = re.search('片头',cols_text[index])
        if pd_titles != None:
            cols_text[index] = ''
        pd_ends = re.search('片尾',cols_text[index])
        if pd_ends != None:
            cols_text[index] = ''

        #去掉近期
        pd_recent = re.search('近期',cols_text[index])
        if pd_recent != None:
            cols_text[index] = ''

        # 去掉金曲
        pd_music = re.search('金曲',cols_text[index])
        if pd_music != None:
            cols_text[index] = ''
        # 宣传片
        pd_AD = re.search('宣传片',cols_text[index])
        if pd_AD != None:
            cols_text[index] = ''
        # 热播
        pd_hot = re.search('热播',cols_text[index])
        if pd_hot != None:
            cols_text[index] = ''
        # 精彩继续
        pd_jingcai = re.search('精彩继续',cols_text[index])
        if pd_jingcai != None:
            cols_text[index] = ''
        # 节目预告
        pd_jiemu = re.search('预告',cols_text[index])
        if pd_jiemu != None:
            cols_text[index] = ''
        #三套
        pd_santao = re.search('三套',cols_text[index])
        if pd_santao != None:
            cols_text[index] = ''
        #银龄大学
        pd_yinling = re.search('银龄大学',cols_text[index])
        if pd_yinling != None:
            cols_text[index] = '银龄大学'

        #去乱七八糟
        for list in lists:
            pd_list = re.search(list,cols_text[index])
            if pd_list != None:
                cols_text[index] = ''

    #去重复
    for index1 in range(len(cols_text)):
        index_qc = index1+1
        if index_qc == len(cols_text): break
        if cols_text[index1] == cols_text[index_qc]:
            cols_text[index_qc] = ''

    # 直播潍坊周三1
    # 广告7a
    # 直播潍坊周三2
    # 广告7b
    # 直播潍坊周三3
    # 广告8
    for ind in range(len(cols_text)):
        if ind + 2 == len(cols_text): break
        if cols_text[ind] == '广告' or cols_text[ind] == '公益广告': continue
        if cols_text[ind] == cols_text[ind+2]:
           cols_text[ind+2] = ''
           for i in range(ind,ind+20):
               if i + 2 == len(cols_text): break
               if cols_text[ind] == cols_text[i + 2]:
                  cols_text[i + 2] = ''



    for index11 in range(len(cols_text)):
        if index11 == len(cols_text)+1: break
        if cols_text[index11] == '广告' or cols_text[index11] == '公益广告':
            # flag.append(index)
            # kbjs = index11 + 1
            for kbjs in range(index11 + 1,len(cols_text)):
                if cols_text[kbjs] != '':
                    # continue
                    if cols_text[kbjs] == '广告' or cols_text[kbjs] == '公益广告':
                        # cols_text[index11] = ''
                        cols_text[kbjs] = ''
                    else :
                        break


    #去空白数据

    for index3 in range(len(cols_text)):
        if cols_text[index3] != '':
            cols_text_s.append(cols_text[index3])
            cols_time_s.append(cols_time[index3])

    #
    # print(cols_text)
    # print(cols_text_s)


def writer_excel(first_name,second_name):

    long = len(cols_text_s)
    wb = openpyxl.Workbook()  # 创建工作簿
    # sheet1 = wb.create_sheet()
    # sheetname = wb.sheetnames
    # sheet1 = wb.get_sheet_by_name(sheetname[0])
    sheet1 = wb.worksheets[0]
    # 生成后续
    sheet1.cell(row=1, column=1,value = '起始时间')
    sheet1.cell(row=1, column=2,value = '节目名称')

    for jkey in range(0,long):
        sheet1.cell(row=jkey+2,column=1).value = cols_time_s[jkey]
        sheet1.cell(row=jkey+2,column=2).value = cols_text_s[jkey]

    # url = first_name
    # 截取选择的文件明生成新的csv格式文件名
    # m = re.findall('[^\\/:*?"<>|\r\n]+$', first_name)
    # p = re.compile(r'.xls')
    # csv_name = p.split(m[0])[0]


    csv_name = second_name
    # url = r'C:\Users\hasee\Desktop'+'\\'+ csv_name +'.xlsx'
    # 首先 生成 .xlsx格式的文件
    url = '.\\' + csv_name + '.xlsx'
    # wb.save(r'C:\Users\hasee\Desktop\chatPy.xlsx')  # 保存文件
    wb.save(url)  # 保存文件

    sfzh = xlsx_to_csv(url,csv_name)   #转成csv文件
    if sfzh == 1:
        os.remove(url)

    cols_time.clear()
    cols_text.clear()
    cols_time_s.clear()
    cols_text_s.clear()

    return

# xlsx转csv
def xlsx_to_csv(url,csv_name):
    # workbook = xlrd.open_workbook(r'.\3.xlsx')
    workbook = xlrd.open_workbook(url)

    table = workbook.sheet_by_index(0)
    with codecs.open('.\\' + csv_name + '.csv', 'w', encoding='gb18030') as f:
        write = csv.writer(f)
        for row_num in range(table.nrows):
            row_value = table.row_values(row_num)
            write.writerow(row_value)
    return 1
# if __name__ == '__main__':
#     # 读取Excel
#     read_excel();
#     writer_excel();
#     print ('读取成功')

def initiate(first_name,second_name):
    # 读取Excel
    read_excel(first_name);
    writer_excel(first_name,second_name);
    print ('读取成功')

    return 1




class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(800, 500)
        self.wj = QtWidgets.QLabel(Form)
        self.wj.setGeometry(QtCore.QRect(40, 40, 100, 40))
        self.wj.setObjectName("wj")
        self.wj.setFont(QFont("宋体", 10,))
        self.mz = QtWidgets.QLabel(Form)
        self.mz.setGeometry(QtCore.QRect(40, 100, 100, 40))
        self.mz.setObjectName("mz")
        self.mz.setFont(QFont("宋体", 10,))
        self.lineEdit = QtWidgets.QLineEdit(Form)
        self.lineEdit.setGeometry(QtCore.QRect(130, 40, 500, 40))
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit.setReadOnly(True)
        self.lineEdit_2 = QtWidgets.QLineEdit(Form)
        self.lineEdit_2.setGeometry(QtCore.QRect(130, 100, 500, 40))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_2.setReadOnly(True)  # 设置为只读，即可以在代码中向textEdit里面输入，但不能从界面上输入,没有这行代码即可以从界面输入

        self.xzan = QtWidgets.QPushButton(Form)
        self.xzan.setGeometry(QtCore.QRect(640, 40, 100, 40))
        self.xzan.setObjectName("xzan")
        self.tj = QtWidgets.QPushButton(Form)
        self.tj.setGeometry(QtCore.QRect(640, 390, 100, 40))
        self.tj.setObjectName("tj")
        self.gc = QtWidgets.QTextEdit(Form)
        self.gc.setGeometry(QtCore.QRect(130, 170, 500, 260))
        self.gc.setObjectName("gc")
        self.gc.setFocusPolicy(QtCore.Qt.NoFocus)
        self.gc.setFontPointSize(16);

        self.ts = QtWidgets.QLabel(Form)
        self.ts.setGeometry(QtCore.QRect(20, 220, 100, 200))
        self.ts.setObjectName("ts")
        self.ts.setFont(QFont("宋体",12,QFont.Bold))
        # self.gc.setColor(QPalette.WindowText, red)  # 设置字体颜色


        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "节目单生成器"))
        self.wj.setText(_translate("Form", "选择文件"))
        self.mz.setText(_translate("Form", "生成文件"))
        self.xzan.setText(_translate("Form", "选择"))
        self.tj.setText(_translate("Form", "提交"))
        self.ts.setText(_translate("Form", "提示：\n\n只能选择\n\n*.xls\n\n格式文件"))





class MyMainForm(QMainWindow, Ui_Form):

    Filename_Paths = []

    def __init__(self, parent=None):

        super(MyMainForm, self).__init__(parent)
        self.setupUi(self)
        self.xzan.clicked.connect(self.openFile)
        self.tj.clicked.connect(self.tj_File)


    def openFile(self):

        get_filename_path, ok = QFileDialog.getOpenFileNames(self,
                                    "选取文件",
                                   # "C:/",
                                    "",
                                    "表格(*.xls)")
        if ok:
            self.lineEdit.setText(str(get_filename_path))
            self.gc.clear()
            global Filename_Paths
            Filename_Paths = get_filename_path




    def tj_File(self):
        self.tj.setEnabled(False)
        for Path in Filename_Paths:
            first_name = Path

            # first_name = self.lineEdit.text()
            # second_name = self.lineEdit_2.text()
            if first_name == '':
                QMessageBox.warning(self, "警告对话框", "请选择文件",)
                self.tj.setEnabled(False)
                return
            # 截取 选择的文件名
            m = re.findall('[^\\/:*?"<>|\r\n]+$', first_name)
            p = re.compile(r'.xls')
            second_name = p.split(m[0])[0]

            # if second_name == '':
            #     QMessageBox.warning(self, "警告对话框", "请命名生成文件", )
            #     return

            self.lineEdit_2.setText(second_name)
            self.gc.insertPlainText('1.选择文件：\n' + '  ' + first_name + '\n')
            self.gc.insertPlainText('2.将生成文件：\n' + '  ' + second_name + '.csv\n')
            self.gc.insertPlainText('3.正在生成请稍后……\n')
            self.gc.insertPlainText('……\n')

            # print(first_name)
            ret_value = initiate(first_name, second_name)
            if ret_value == 1:
                self.tj.setEnabled(True)
                self.lineEdit.clear()
                # self.lineEdit_2.clear()
                QMessageBox.information(self,'','已生成，请在该程序同目录查看   ')

if __name__ == "__main__":
    #固定的，PyQt5程序都需要QApplication对象。sys.argv是命令行参数列表，确保程序可以双击运行
    app = QApplication(sys.argv)
    #初始化
    myWin = MyMainForm()
    #将窗口控件显示在屏幕上
    myWin.show()
    #程序运行，sys.exit方法确保程序完整退出。
    sys.exit(app.exec_())















